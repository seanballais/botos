from openpyxl import Workbook
from openpyxl.styles import (
    Alignment, Font
)
from openpyxl.utils import get_column_letter
from openpyxl.writer.excel import save_virtual_workbook

from django.contrib import messages
from django.db.models import (
    Count, Q
)
from django.http import HttpResponse
from django.shortcuts import redirect
from django.urls import reverse
from django.utils.decorators import method_decorator
from django.views import View

from core.decorators import (
    login_required, user_passes_test
)
from core.models import (
    Batch, Candidate, CandidateParty, CandidatePosition,
    Election, Section, UserType, Vote, VoterProfile
)


@method_decorator(
    login_required(
        login_url='/',
        next='/admin/results'
    ),
    name='dispatch',
)
@method_decorator(
    user_passes_test(
        lambda u: u.type == UserType.ADMIN,
        login_url='/',
        next='',
        redirect_field_name=None
    ),
    name='dispatch',
)
class ResultsExporterView(View):
    """
    View that exports results. Currently, it only exports to an XLSX file.

    This subview may only process requests from admin users. Other users will
    be redirected to '/'.

    The view may accept a URL parameter, election, to return the results of a
    specific election. Invalid values for the election parameter will result in
    the view returning an error message.

    View URL: 'admin/results/export'
    """
    def get(self, request):
        election_id = request.GET.get('election', None)
        filename = ''
        if election_id:
            try:
                election_id = int(election_id)
            except ValueError:
                # The election id is a non-integer.
                messages.error(
                    request,
                    'You specified a non-integer election ID.'
                )
                return redirect(request.META.get('HTTP_REFERER', '/'))
            else:
                try:
                    election_name = Election.objects.get(id=election_id).name
                except Election.DoesNotExist:
                    messages.error(
                        request,
                        'You specified an ID for a non-existent election.'
                    )
                    return redirect(request.META.get('HTTP_REFERER', '/'))
                else:
                    filename = '{} Results.xlsx'.format(election_name)
        else:
            filename = 'Election Results.xlsx'

        wb = self._generate_xlsx_file(election_id)
        content_type = (
            'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        content_disposition = 'attachment; filename="{}"'.format(filename)
        response = HttpResponse(
            content=save_virtual_workbook(wb),
            content_type=content_type
        )
        response['Content-Disposition'] = content_disposition

        return response

    def _generate_xlsx_file(self, election_id):
        elections = Election.objects.all()
        if election_id:
            elections = [ elections.get(id=election_id) ]

        num_worksheets = len(elections)
        wb = Workbook()

        # Create the necessary worksheets.
        for i in range(num_worksheets - 1):
            # We subtract by one since we have at least one worksheet in a
            # workbook upon creation.
            wb.create_sheet("")

        for sheet_idx, election in enumerate(elections):
            ws = wb.worksheets[sheet_idx]

            ws.title = election.name

            # Set up the sheet header title.
            num_columns = VoterProfile.objects                          \
                                      .filter(batch__election=election) \
                                      .order_by()                       \
                                      .distinct('section')              \
                                      .count()

            # +2 since we have a column for the candidates and another column
            # for the total votes.
            num_columns += 2

            ws.merge_cells(
                start_row=1,
                start_column=1,
                end_row=1,
                end_column=num_columns
            )
            ws.cell(1, 1).value = '{} Results'.format(election.name)
            ws.cell(1, 1).alignment = Alignment(
                horizontal='center',
                vertical='center'
            )

            ws.merge_cells(
                start_row=2,
                start_column=2,
                end_row=2,
                end_column=num_columns
            )
            ws.cell(2, 2).value = 'Number of Votes'
            ws.cell(2, 2).alignment = Alignment(horizontal='center')

            # Set up the candidate block and total votes block.
            ws.merge_cells('A2:A4')
            ws.cell(2, 1).value = 'Candidates'
            ws.cell(2, 1).alignment = Alignment(
                horizontal='center',
                vertical='center'
            )

            ws.merge_cells(
                start_row=3, 
                start_column=num_columns,
                end_row=4,
                end_column=num_columns
            )
            ws.cell(3, num_columns).value = 'Total Votes'
            ws.cell(3, num_columns).alignment = Alignment(
                horizontal='center',
                vertical='center'
            )
            ws.column_dimensions[get_column_letter(num_columns)].width = 14

            # Set up batches and sections.
            batches = Batch.objects.filter(
                election=election,
                voter_profiles__isnull=False
            ).distinct()

            curr_batch_col = 2
            for batch in batches:
                sections = Section.objects                             \
                                  .filter(voter_profiles__batch=batch) \
                                  .distinct()

                if len(sections) > 0:
                    ws.merge_cells(
                        start_row=3,
                        start_column=curr_batch_col,
                        end_row=3,
                        # We need a -1 here, because, otherwise, we'll invade
                        # one cell of an adjacent batch cell.
                        end_column=curr_batch_col + (len(sections) - 1)
                    )

                ws.cell(3, curr_batch_col).value = str(batch)
                ws.cell(3, curr_batch_col).alignment = Alignment(
                    horizontal='center'
                )

                for section_idx, section in enumerate(sections):
                    section_col = curr_batch_col + section_idx
                    ws.cell(4, section_col).value = str(section)
                    ws.cell(4, section_col).alignment = Alignment(
                        horizontal='center'
                    )
                    col_letter = get_column_letter(section_col)
                    col_width = len(str(section)) + 4
                    ws.column_dimensions[col_letter].width = col_width

                curr_batch_col += len(sections)

            # Set up the candidate column.
            len_longest_cand_name = 0
            positions = CandidatePosition.objects.filter(election=election)
            position_start_row = 5
            for position_idx, position in enumerate(positions):
                ws.cell(position_start_row, 1).value = str(position)
                ws.cell(position_start_row, 1).font = Font(bold=True)

                parties = CandidateParty.objects.filter(election=election)

                # Distance of the current empty cell to the current position
                # cell.
                dist_to_curr_pos_cell = 0
                for party in parties:
                    dist_to_curr_pos_cell += 1
                    party_pos = position_start_row + dist_to_curr_pos_cell
                    ws.cell(party_pos, 1).value = str(party)
                    ws.cell(party_pos, 1).font = Font(italic=True)

                    candidates = Candidate.objects.filter(
                        party=party,
                        position=position,
                        election=election
                    )
                    candidates = candidates.annotate(
                        total_votes=Count('votes')
                    )
                    if len(candidates) > 0:
                        candidates_enum = enumerate(candidates, 1)
                        for candidate_idx, candidate in candidates_enum:
                            self._write_candidate_votes(
                                ws, election, candidate,
                                candidate_idx, party_pos, num_columns
                            )

                            if len(str(candidate)) > len_longest_cand_name:
                                len_longest_cand_name = len(str(candidate))

                        dist_to_curr_pos_cell += len(candidates)
                    else:
                        candidate_row = party_pos + 1
                        
                        self._write_no_candidate_cells(
                            ws,
                            election,
                            candidate_row,
                            num_columns
                        )

                        dist_to_curr_pos_cell += 1

                position_start_row += dist_to_curr_pos_cell + 1

            # +12 for spacing.
            ws.column_dimensions['A'].width = len_longest_cand_name + 12

            sheet_idx += 1

        return wb

    def _write_candidate_votes(self,
                               ws,
                               election,
                               candidate,
                               candidate_idx,
                               party_pos,
                               num_columns):
        candidate_row = party_pos + candidate_idx
        ws.cell(candidate_row, 1).value = str(candidate)

        batches = Batch.objects.filter(election=election)
        curr_section_col = 2
        for batch in batches:
            sections = Section.objects                             \
                              .filter(voter_profiles__batch=batch) \
                              .distinct()
            for section in sections:
                num_votes = Vote.objects.filter(
                    candidate=candidate,
                    election=election,
                    user__voter_profile__section=section
                ).count()

                ws.cell(candidate_row, curr_section_col).value = num_votes

                curr_section_col += 1

        total_votes_cell = ws.cell(candidate_row, num_columns)
        total_votes_cell.value = candidate.total_votes
        total_votes_cell.alignment = Alignment(horizontal='right')

    def _write_no_candidate_cells(self,
                                  ws,
                                  election,
                                  candidate_row,
                                  num_columns):
        total_votes_cell = ws.cell(candidate_row, num_columns)

        batches = Batch.objects.filter(election=election)
        curr_section_col = 2
        for batch in batches:
            sections = Section.objects                             \
                              .filter(voter_profiles__batch=batch) \
                              .distinct()
            for section in sections:
                ws.cell(candidate_row, curr_section_col).value = 'N/A'
                ws.cell(candidate_row, curr_section_col).alignment = Alignment(
                    horizontal='right'
                )

                curr_section_col += 1

        ws.cell(candidate_row, 1).value = 'None'
        total_votes_cell.value = 'N/A'
        total_votes_cell.alignment = Alignment(horizontal='right')
