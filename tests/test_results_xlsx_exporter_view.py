import io
import zipfile

from django.test import (
    Client, TestCase
)
from django.urls import reverse
import openpyxl

from core.models import (
    User, Batch, Section, Election, Candidate, CandidateParty,
    CandidatePosition, Vote, VoterProfile, Setting, UserType
)


class ResultsExporter(TestCase):
    """
    Tests the results xlsx exporter view.

    This subview may only process requests from logged in admin users. Other
    users will be redirected to '/'. This will also only accept GET requests.

    GET requests may have an election`parameter whose value must be the id
    of an election. The lack of an election parameter will result in the
    results of all elections to be exported, with each election having its
    own worksheet. Other URL parameters will be ignored.

    View URL: '/results/export'
    """
    @classmethod
    def setUpTestData(cls):
        batch_num = 0
        section_num = 0
        voter_num = 0
        party_num = 0
        position_num = 0
        candidate_num = 0
        num_elections = 2
        voters = list()
        for i in range(num_elections):
            election = Election.objects.create(name='Election {}'.format(i))

            num_batches = 2
            for j in range(num_batches):
                batch = Batch.objects.create(year=batch_num, election=election)
                batch_num += 1

                num_sections = 2 if j == 0 else 1
                for k in range(num_sections):
                    section = Section.objects.create(
                        section_name='Section {}'.format(section_num)
                    )
                    section_num += 1

                    voter = User.objects.create(
                        username='user{}'.format(voter_num),
                        first_name=str(voter_num),
                        last_name=str(voter_num),
                        type=UserType.VOTER
                    )
                    voter.set_password('voter')
                    voter.save()
                    voter_num += 1

                    VoterProfile.objects.create(
                        user=voter,
                        batch=batch,
                        section=section
                    )

                    voters.append(voter)

            num_positions = 3
            for i in range(num_positions):
                if str(election.name) not in positions:
                    positions[str(election.name)] = list()

                position = CandidatePosition.objects.create(
                    position_name='Position {}'.format(position_num)
                )

                position_num += 1

                positions[str(election.name)].append(position)

            num_parties = 2
            for j in range(num_parties):
                party = CandidateParty.objects.create(
                    party_name='Party {}'.format(party_num),
                    election=election
                )
                party_num += 1

                num_positions = 3
                for k in range(num_positions):
                    position = positions[str(election.name)][k]

                    candidate = Candidate.objects.create(
                        user=voters[candidate_num],
                        party=party,
                        position=position,
                        election=election
                    )

                    Vote.objects.create(
                        user=voters[candidate_num],
                        candidate=candidate,
                        election=election
                    )

                    candidate_num += 1

        # Let's give one candidate an additional vote to really make sure that
        # we all got the correct number of votes.
        Vote.objects.create(
            user=voters[0],
            # NOTE: The voter in voter[1] is a Position 1 candidate of
            #       Party 1, where the voter in voter[0] is a member.
            candidate=Candidate.objects.get(user=voters[1]),
            election=Election.objects.get(name='Election 0')
        )

        _admin = User.objects.create(username='admin', type=UserType.ADMIN)
        _admin.set_password('root')
        _admin.save()

    def setUp(self):
        self.client.login(username='admin', password='root')

    def test_anonymous_get_requests_redirected_to_index(self):
        self.client.logout()

        response = self.client.get(reverse('results-export'), follow=True)
        self.assertRedirects(response, reverse('index'))

    def test_voter_get_requests_redirected_to_index(self):
        self.client.login(username='user0', password='voter')

        response = self.client.get(reverse('results-export'), follow=True)
        self.assertRedirects(response, reverse('index'))

    def test_get_all_elections_xlsx(self):
        response = self.client.get(reverse('results-export'))

        self.assertEqual(response.status_code, 200)

        xlsx_file = zipfile.ZipFile(io.BytesIO(response.content), 'r')
        self.assertIn('Election Results.xlsx', xlsx_file.namelist())

        wb = openpyxl.load_workbook(xlsx_file)

        self.assertEqual(len(wb.worksheets), 2)

        # Check first worksheet.
        ws = wb.worksheets[0]

        self.assertEqual(wb.sheetnames[0], 'Election 0')

        row_count = ws.max_row
        col_count = ws.max_col
        self.assertEqual(row_count, 13)
        self.assertEqual(col_count, 5)

        # Check candidate column.
        self.assertEqual(str(ws.cell(2, 1).value), 'Candidates')
        self.assertEqual(str(ws.cell(5, 1).value), 'Position 0')
        self.assertEqual(str(ws.cell(6, 1).value), '0, 0')
        self.assertEqual(str(ws.cell(7, 1).value), '3, 3')
        self.assertEqual(str(ws.cell(8, 1).value), 'Position 1')
        self.assertEqual(str(ws.cell(9, 1).value), '1, 1')
        self.assertEqual(str(ws.cell(10, 1).value), '4, 4')
        self.assertEqual(str(ws.cell(11, 1).value), 'Position 2')
        self.assertEqual(str(ws.cell(12, 1).value), '2, 2')
        self.assertEqual(str(ws.cell(13, 1).value), '5, 5')

    def test_get_election0_xlsx(self):
        pass
