import io

import openpyxl

from django.test import (
    Client, TestCase
)
from django.urls import reverse

from core.models import (
    User, Batch, Section, Election, Candidate, CandidateParty,
    CandidatePosition, Vote, VoterProfile, Setting, UserType
)


class ResultsExporter(TestCase):
    """
    Tests the results xlsx exporter view.

    This subview may only process requests from logged in admin users. Other
    users will be redirected to '/'. This will also only accept GET requests.

    GET requests may have an election`parameter whose value must be the id
    of an election. The lack of an election parameter will result in the
    results of all elections to be exported, with each election having its
    own worksheet. Other URL parameters will be ignored. Invalid election
    parameter values, e.g. non-existent election IDs and non-integer parameters,
    will return an error message.

    View URL: '/results/export'
    """
    @classmethod
    def setUpTestData(cls):
        batch_num = 0
        section_num = 0
        voter_num = 0
        party_num = 0
        position_num = 0
        candidate_num = 0
        num_elections = 2
        voters = list()
        positions = dict()
        for i in range(num_elections):
            election = Election.objects.create(name='Election {}'.format(i))
            positions[str(election.name)] = list()

            num_batches = 2
            for j in range(num_batches):
                batch = Batch.objects.create(year=batch_num, election=election)
                batch_num += 1

                num_sections = 2 if j == 0 else 1
                for k in range(num_sections):
                    section = Section.objects.create(
                        section_name=str(section_num)
                    )
                    section_num += 1

                    num_students = 2
                    for l in range(num_students):
                        voter = User.objects.create(
                            username='user{}'.format(voter_num),
                            first_name=str(voter_num),
                            last_name=str(voter_num),
                            type=UserType.VOTER
                        )
                        voter.set_password('voter')
                        voter.save()
                        voter_num += 1

                        VoterProfile.objects.create(
                            user=voter,
                            batch=batch,
                            section=section
                        )

                        voters.append(voter)

            num_positions = 3
            for i in range(num_positions):
                position = CandidatePosition.objects.create(
                    position_name='Position {}'.format(position_num),
                    election=election
                )

                positions[str(election.name)].append(position)

                position_num += 1

            num_parties = 3
            for j in range(num_parties):
                party = CandidateParty.objects.create(
                    party_name='Party {}'.format(party_num),
                    election=election
                )
                party_num += 1

                if j != 2:  # Let every third party have no candidates.
                    num_positions = 3
                    for k in range(num_positions):
                        position = positions[str(election.name)][k]

                        candidate = Candidate.objects.create(
                            user=voters[candidate_num],
                            party=party,
                            position=position,
                            election=election
                        )

                        Vote.objects.create(
                            user=voters[candidate_num],
                            candidate=candidate,
                            election=election
                        )

                        candidate_num += 1

        # Let's give one candidate an additional vote to really make sure that
        # we all got the correct number of votes.
        Vote.objects.create(
            user=voters[0],
            # NOTE: The voter in voter[1] is a Position 1 candidate of
            #       Party 1, where the voter in voter[0] is a member.
            candidate=Candidate.objects.get(user=voters[1]),
            election=Election.objects.get(name='Election 0')
        )

        _admin = User.objects.create(username='admin', type=UserType.ADMIN)
        _admin.set_password('root')
        _admin.save()

    def setUp(self):
        self.client.login(username='admin', password='root')

    def test_anonymous_get_requests_redirected_to_index(self):
        self.client.logout()

        response = self.client.get(reverse('results-export'), follow=True)
        self.assertRedirects(response, '/?next=%2Fadmin%2Fresults')

    def test_voter_get_requests_redirected_to_index(self):
        self.client.logout()
        self.client.login(username='user0', password='voter')

        response = self.client.get(reverse('results-export'), follow=True)
        self.assertRedirects(response, reverse('index'))

    def test_get_all_elections_xlsx(self):
        response = self.client.get(reverse('results-export'))

        self.assertEqual(response.status_code, 200)

        self.assertEqual(
            response['Content-Disposition'],
            'attachment; filename="Election Results.xlsx"'
        )

        wb = openpyxl.load_workbook(io.BytesIO(response.content))

        self.assertEqual(len(wb.worksheets), 2)

        # Check first worksheet.
        ws = wb.worksheets[0]

        self.assertEqual(wb.sheetnames[0], 'Election 0')

        row_count = ws.max_row
        col_count = ws.max_column
        self.assertEqual(row_count, 25)
        self.assertEqual(col_count, 5)

        self.assertEqual(str(ws.cell(1, 1).value), 'Election 0 Results')

        self.assertEqual(str(ws.cell(2, 1).value), 'Candidates')

        cellContents = [
            'Position 0',
            'Party 0',
            '0, 0',
            'Party 1',
            '3, 3',
            'Party 2',
            'None',
            'Position 1',
            'Party 0',
            '1, 1',
            'Party 1',
            '4, 4',
            'Party 2',
            'None',
            'Position 2',
            'Party 0',
            '2, 2',
            'Party 1',
            '5, 5',
            'Party 2',
            'None'
        ]
        for cellIndex, content in enumerate(cellContents, 5):
            self.assertEqual(str(ws.cell(cellIndex, 1).value), content) 

        self.assertEqual(str(ws.cell(2, 2).value), 'Number of Votes')

        self.assertEqual(str(ws.cell(3, 2).value), '0')

        self.assertEqual(str(ws.cell(4, 2).value), '0') # Section

        self.assertEqual(str(ws.cell(7, 2).value), '1')
        self.assertEqual(str(ws.cell(9, 2).value), '0')
        self.assertEqual(str(ws.cell(11, 2).value), 'N/A')
        self.assertEqual(str(ws.cell(14, 2).value), '2')
        self.assertEqual(str(ws.cell(16, 2).value), '0')
        self.assertEqual(str(ws.cell(18, 2).value), 'N/A')
        self.assertEqual(str(ws.cell(21, 2).value), '0')
        self.assertEqual(str(ws.cell(23, 2).value), '0')
        self.assertEqual(str(ws.cell(25, 2).value), 'N/A')

        self.assertEqual(str(ws.cell(4, 3).value), '1') # Section

        self.assertEqual(str(ws.cell(7, 3).value), '0')
        self.assertEqual(str(ws.cell(9, 3).value), '1')
        self.assertEqual(str(ws.cell(11, 2).value), 'N/A')
        self.assertEqual(str(ws.cell(14, 3).value), '0')
        self.assertEqual(str(ws.cell(16, 3).value), '0')
        self.assertEqual(str(ws.cell(18, 2).value), 'N/A')
        self.assertEqual(str(ws.cell(21, 3).value), '1')
        self.assertEqual(str(ws.cell(23, 3).value), '0')
        self.assertEqual(str(ws.cell(25, 2).value), 'N/A')

        self.assertEqual(str(ws.cell(3, 4).value), '1')

        self.assertEqual(str(ws.cell(4, 4).value), '2') # Section

        self.assertEqual(str(ws.cell(7, 4).value), '0')
        self.assertEqual(str(ws.cell(9, 4).value), '0')
        self.assertEqual(str(ws.cell(11, 2).value), 'N/A')
        self.assertEqual(str(ws.cell(14, 4).value), '0')
        self.assertEqual(str(ws.cell(16, 4).value), '1')
        self.assertEqual(str(ws.cell(18, 2).value), 'N/A')
        self.assertEqual(str(ws.cell(21, 4).value), '0')
        self.assertEqual(str(ws.cell(23, 4).value), '1')
        self.assertEqual(str(ws.cell(25, 2).value), 'N/A')

        self.assertEqual(str(ws.cell(3, 5).value), 'Total Votes')
        self.assertEqual(str(ws.cell(7, 5).value), '1')
        self.assertEqual(str(ws.cell(9, 5).value), '1')
        self.assertEqual(str(ws.cell(11, 2).value), 'N/A')
        self.assertEqual(str(ws.cell(14, 5).value), '2')
        self.assertEqual(str(ws.cell(16, 5).value), '1')
        self.assertEqual(str(ws.cell(18, 2).value), 'N/A')
        self.assertEqual(str(ws.cell(21, 5).value), '1')
        self.assertEqual(str(ws.cell(23, 5).value), '1')
        self.assertEqual(str(ws.cell(25, 2).value), 'N/A')

        # Check second worksheet.
        ws = wb.worksheets[1]

        self.assertEqual(wb.sheetnames[1], 'Election 1')

        row_count = ws.max_row
        col_count = ws.max_column
        self.assertEqual(row_count, 25)
        self.assertEqual(col_count, 5)

        self.assertEqual(str(ws.cell(1, 1).value), 'Election 1 Results')

        self.assertEqual(str(ws.cell(2, 1).value), 'Candidates')

        self.assertEqual(str(ws.cell(2, 1).value), 'Candidates')

        cellContents = [
            'Position 3',
            'Party 3',
            '6, 6',
            'Party 4',
            '9, 9',
            'Party 5',
            'None',
            'Position 4',
            'Party 3',
            '7, 7',
            'Party 4',
            '10, 10',
            'Party 5',
            'None',
            'Position 5',
            'Party 3',
            '8, 8',
            'Party 4',
            '11, 11',
            'Party 5',
            'None'
        ]
        for cellIndex, content in enumerate(cellContents, 5):
            self.assertEqual(str(ws.cell(cellIndex, 1).value), content) 

        self.assertEqual(str(ws.cell(2, 2).value), 'Number of Votes')

        self.assertEqual(str(ws.cell(3, 2).value), '2')

        self.assertEqual(str(ws.cell(4, 2).value), '3') # Section

        self.assertEqual(str(ws.cell(7, 2).value), '1')
        self.assertEqual(str(ws.cell(9, 2).value), '0')
        self.assertEqual(str(ws.cell(11, 2).value), 'N/A')
        self.assertEqual(str(ws.cell(14, 2).value), '1')
        self.assertEqual(str(ws.cell(16, 2).value), '0')
        self.assertEqual(str(ws.cell(18, 2).value), 'N/A')
        self.assertEqual(str(ws.cell(21, 2).value), '0')
        self.assertEqual(str(ws.cell(23, 2).value), '0')
        self.assertEqual(str(ws.cell(25, 2).value), 'N/A')

        self.assertEqual(str(ws.cell(4, 3).value), '4') # Section

        self.assertEqual(str(ws.cell(7, 3).value), '0')
        self.assertEqual(str(ws.cell(9, 3).value), '1')
        self.assertEqual(str(ws.cell(11, 2).value), 'N/A')
        self.assertEqual(str(ws.cell(14, 3).value), '0')
        self.assertEqual(str(ws.cell(16, 3).value), '0')
        self.assertEqual(str(ws.cell(18, 2).value), 'N/A')
        self.assertEqual(str(ws.cell(21, 3).value), '1')
        self.assertEqual(str(ws.cell(23, 3).value), '0')
        self.assertEqual(str(ws.cell(25, 2).value), 'N/A')

        self.assertEqual(str(ws.cell(3, 4).value), '3')

        self.assertEqual(str(ws.cell(4, 4).value), '5') # Section

        self.assertEqual(str(ws.cell(7, 4).value), '0')
        self.assertEqual(str(ws.cell(9, 4).value), '0')
        self.assertEqual(str(ws.cell(11, 2).value), 'N/A')
        self.assertEqual(str(ws.cell(14, 4).value), '0')
        self.assertEqual(str(ws.cell(16, 4).value), '1')
        self.assertEqual(str(ws.cell(18, 2).value), 'N/A')
        self.assertEqual(str(ws.cell(21, 4).value), '0')
        self.assertEqual(str(ws.cell(23, 4).value), '1')
        self.assertEqual(str(ws.cell(25, 2).value), 'N/A')

        self.assertEqual(str(ws.cell(3, 5).value), 'Total Votes')
        self.assertEqual(str(ws.cell(7, 5).value), '1')
        self.assertEqual(str(ws.cell(9, 5).value), '1')
        self.assertEqual(str(ws.cell(11, 2).value), 'N/A')
        self.assertEqual(str(ws.cell(14, 5).value), '1')
        self.assertEqual(str(ws.cell(16, 5).value), '1')
        self.assertEqual(str(ws.cell(18, 2).value), 'N/A')
        self.assertEqual(str(ws.cell(21, 5).value), '1')
        self.assertEqual(str(ws.cell(23, 5).value), '1')
        self.assertEqual(str(ws.cell(25, 2).value), 'N/A')

    def test_get_election0_xlsx(self):
        response = self.client.get(
            reverse('results-export'),
            { 'election': str(Election.objects.get(name='Election 0').id) }
        )

        self.assertEqual(response.status_code, 200)        

        self.assertEqual(
            response['Content-Disposition'],
            'attachment; filename="Election 0 Results.xlsx"'
        )

        wb = openpyxl.load_workbook(io.BytesIO(response.content))

        self.assertEqual(len(wb.worksheets), 1)

        # Check first worksheet.
        ws = wb.worksheets[0]

        self.assertEqual(wb.sheetnames[0], 'Election 0')

        row_count = ws.max_row
        col_count = ws.max_column
        self.assertEqual(row_count, 25)
        self.assertEqual(col_count, 5)

        self.assertEqual(str(ws.cell(1, 1).value), 'Election 0 Results')

        self.assertEqual(str(ws.cell(2, 1).value), 'Candidates')

        cellContents = [
            'Position 0',
            'Party 0',
            '0, 0',
            'Party 1',
            '3, 3',
            'Party 2',
            'None',
            'Position 1',
            'Party 0',
            '1, 1',
            'Party 1',
            '4, 4',
            'Party 2',
            'None',
            'Position 2',
            'Party 0',
            '2, 2',
            'Party 1',
            '5, 5',
            'Party 2',
            'None'
        ]
        for cellIndex, content in enumerate(cellContents, 5):
            self.assertEqual(str(ws.cell(cellIndex, 1).value), content) 

        self.assertEqual(str(ws.cell(2, 2).value), 'Number of Votes')

        self.assertEqual(str(ws.cell(3, 2).value), '0')

        self.assertEqual(str(ws.cell(4, 2).value), '0') # Section

        self.assertEqual(str(ws.cell(7, 2).value), '1')
        self.assertEqual(str(ws.cell(9, 2).value), '0')
        self.assertEqual(str(ws.cell(11, 2).value), 'N/A')
        self.assertEqual(str(ws.cell(14, 2).value), '2')
        self.assertEqual(str(ws.cell(16, 2).value), '0')
        self.assertEqual(str(ws.cell(18, 2).value), 'N/A')
        self.assertEqual(str(ws.cell(21, 2).value), '0')
        self.assertEqual(str(ws.cell(23, 2).value), '0')
        self.assertEqual(str(ws.cell(25, 2).value), 'N/A')

        self.assertEqual(str(ws.cell(4, 3).value), '1') # Section

        self.assertEqual(str(ws.cell(7, 3).value), '0')
        self.assertEqual(str(ws.cell(9, 3).value), '1')
        self.assertEqual(str(ws.cell(11, 2).value), 'N/A')
        self.assertEqual(str(ws.cell(14, 3).value), '0')
        self.assertEqual(str(ws.cell(16, 3).value), '0')
        self.assertEqual(str(ws.cell(18, 2).value), 'N/A')
        self.assertEqual(str(ws.cell(21, 3).value), '1')
        self.assertEqual(str(ws.cell(23, 3).value), '0')
        self.assertEqual(str(ws.cell(25, 2).value), 'N/A')

        self.assertEqual(str(ws.cell(3, 4).value), '1')

        self.assertEqual(str(ws.cell(4, 4).value), '2') # Section

        self.assertEqual(str(ws.cell(7, 4).value), '0')
        self.assertEqual(str(ws.cell(9, 4).value), '0')
        self.assertEqual(str(ws.cell(11, 2).value), 'N/A')
        self.assertEqual(str(ws.cell(14, 4).value), '0')
        self.assertEqual(str(ws.cell(16, 4).value), '1')
        self.assertEqual(str(ws.cell(18, 2).value), 'N/A')
        self.assertEqual(str(ws.cell(21, 4).value), '0')
        self.assertEqual(str(ws.cell(23, 4).value), '1')
        self.assertEqual(str(ws.cell(25, 2).value), 'N/A')

        self.assertEqual(str(ws.cell(3, 5).value), 'Total Votes')
        self.assertEqual(str(ws.cell(7, 5).value), '1')
        self.assertEqual(str(ws.cell(9, 5).value), '1')
        self.assertEqual(str(ws.cell(11, 2).value), 'N/A')
        self.assertEqual(str(ws.cell(14, 5).value), '2')
        self.assertEqual(str(ws.cell(16, 5).value), '1')
        self.assertEqual(str(ws.cell(18, 2).value), 'N/A')
        self.assertEqual(str(ws.cell(21, 5).value), '1')
        self.assertEqual(str(ws.cell(23, 5).value), '1')
        self.assertEqual(str(ws.cell(25, 2).value), 'N/A')

    def test_get_with_invalid_election_id_non_existent_election_id(self):
        response = self.client.get(
            reverse('results-export'),
            { 'election': '69' },
            HTTP_REFERER=reverse('results'),
            follow=True
        )

        messages = list(response.context['messages'])
        self.assertEqual(
            messages[0].message,
            'You specified an ID for a non-existent election.'
        )
        self.assertRedirects(response, reverse('results'))

    def test_get_with_invalid_election_id_non_integer_election_id(self):
        response = self.client.get(
            reverse('results-export'),
            { 'election': 'hey' },
            HTTP_REFERER=reverse('results'),
            follow=True
        )

        messages = list(response.context['messages'])
        self.assertEqual(
            messages[0].message,
            'You specified a non-integer election ID.'
        )
        self.assertRedirects(response, reverse('results'))

    def test_ref_get_with_invalid_election_id_non_existent_election_id(self):
        response = self.client.get(
            reverse('results-export'),
            { 'election': '69' },
            HTTP_REFERER=reverse('results'),
            follow=True
        )

        messages = list(response.context['messages'])
        self.assertEqual(
            messages[0].message,
            'You specified an ID for a non-existent election.'
        )
        self.assertRedirects(response, reverse('results'))

    def test_ref_get_with_invalid_election_id_non_integer_election_id(self):
        response = self.client.get(
            reverse('results-export'),
            { 'election': 'hey' },
            HTTP_REFERER=reverse('results'),
            follow=True
        )

        messages = list(response.context['messages'])
        self.assertEqual(
            messages[0].message,
            'You specified a non-integer election ID.'
        )
        self.assertRedirects(response, reverse('results'))
